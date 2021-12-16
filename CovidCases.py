import requests                 
from os import path
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

class Cases:
    def __init__(self, wb, soup, nameOfWb, nameOfWs):   #Constructor for cases class which has a workbook object, soup object, and two stings for name of workbook and worksheet
        self.wb = wb
        self.soup = soup
        self.nameOfWb = nameOfWb
        self.nameOfWs = nameOfWs

    def getWebsite(self, url):      #Takes info from url passes in by argument and stores it in the response which is then converted to a soup object for easy searching
        response = requests.get(url)
        self.soup = BeautifulSoup(response.text, "html.parser")
    
    def getInfo(self):      #Uses soup object to search for day, average cases, and total then returns all three
        day = self.soup.find_all('th', {'class': "svelte-6tbkhx"})[1].text
        cases = self.soup.find_all('td', {'class':"num cases svelte-6tbkhx"})[0].text
        total = self.soup.find_all('td', {'class': 'num cases svelte-6tbkhx'})[2].text

        return day, cases, total
    
    def printInfo(self, day, cases, total):     #Prints the info obtained by getInfo
        print(day +": " + cases + ' cases')
        print("Total Cases: " + total + ' cases')
 
    def outPutInfo(self, day, cases, total):    #Uses the workbook object and info passed by argument to then either create a new workbook and store info or store info in an existing workbook
        if path.exists(self.nameOfWb):
            self.wb = load_workbook(self.nameOfWb) 
            ws = self.wb.active 
            ws.append([day, cases, total])
        else:
            self.wb = Workbook()
            ws = self.wb.active
            ws.title = self.nameOfWs
            ws.append(['Date','Cases','Total'])
            ws.append([day, cases, total])

        self.wb.save(self.nameOfWb)

def main():
    url = "https://www.nytimes.com/interactive/2021/us/hawaii-covid-cases.html"

    case = Cases(None, None, "Covid Cases Hawaii.xlsx", "Covid Cases Hawaii")
    case.getWebsite(url)
    day, cases, total = case.getInfo()
    case.printInfo(day, cases, total)

   # case.outPutInfo(day, cases, total)

if __name__ == '__main__':
    main()
